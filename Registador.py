import time
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from datetime import date
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from openpyxl import load_workbook
from tkinter import *
import threading
from PIL import ImageTk, Image


def mensage(linha):
    global msg
    if msg == "":
        msg = ("ERRO: Linha numero %s nao registada." % linha)
    else:
        msg = msg[0:len(msg)-14] + linha + " nao registada."


def repeat():
    slide2()


def repeat2():
    slide()


def slide():
    for i in range(450):
        Label(root, text=msg, font=('Nunito', 12, 'bold'),
              fg="#DA2326").place(x=2 * i, y=437, anchor="e")
        time.sleep(0.02)
        if i == 300:
            t1 = threading.Thread(target=repeat, args=(i,))
            t1.start()


def slide2():
    for i in range(450):
        Label(root, text=msg, font=('Nunito', 12, 'bold'),
              fg="#DA2326").place(x=2 * i, y=437, anchor="e")
        time.sleep(0.02)
        if i == 300:
            t2 = threading.Thread(target=repeat2, args=(i,))
            t2.start()


def normal():
    numero = 10000
    contador = 0
    global linha
    wb = load_workbook('assets/Normais.xlsx')
    ws = wb.active
    today = date.today()
    today = today.strftime("%d/%m/%Y")
    today = today.split('/')
    website = webdriver.Chrome("assets/chromedriver.exe")
    website.get('https://sinave.min-saude.pt/SINAVE.MIN-SAUDE/login.html')
    elem = website.find_element_by_name("username")
    elem.send_keys('******')
    elem = website.find_element_by_name("password")
    elem.send_keys('******')
    elem = website.find_element_by_name("button")
    elem.click()
    elem = website.find_element_by_name("chooselab")
    elem.click()
    for i in range(numero):
        linha = str(i + 2)
        try:
            if ws['C' + linha].value != "-":
                contador += 1
                notificacao = website.find_element_by_xpath('//*[@id="formTable"]/div/div[1]/div/div[1]/h3')
                notificacao = str(int(notificacao.text)+1)
                website.get('https://sinave.min-saude.pt/SINAVE.MIN-SAUDE/DiseaseNotificationPresc')
                elem = website.find_element_by_name("num_exame_form")
                elem.send_keys(notificacao)
                elem = website.find_element_by_name("presctyperadio")
                elem.click()
                elem = website.find_element_by_name("cod_local_colheita_form")
                elem.send_keys("******")
                elem = website.find_element_by_name("codpostal_local_colheita_form")
                elem.send_keys("******")
                elem = website.find_element_by_id("nextbtn")
                elem.click()
                elem = website.find_element_by_id("patient")
                utente = ws['B' + linha].value
                elem.send_keys(utente)
                elem = website.find_element_by_xpath('//*[@id="patient-box"]/div[1]/section[2]/label/button')
                elem.click()
                nome = website.find_element_by_name('nome_utente_form')
                nome.get_attribute("value")
                morada = website.find_element_by_name('morada_utente_form')
                morada.get_attribute("value")
                nascimento = website.find_element_by_name('utente_data_nasc_form')
                nascimento.get_attribute("value")
                elem = website.find_element_by_id("nextbtn")
                elem.click()
                elem = Select(website.find_element_by_id('diseaseDropList:1'))
                elem.select_by_value('62')
                elem = Select(website.find_element_by_id('productDropList:1'))
                elem.select_by_value('137')
                elem = website.find_element_by_id('data_colheita_form:1')
                elem.send_keys(today[0])
                elem.send_keys(today[1])
                elem.send_keys(today[2])
                horas1 = str(ws['F' + linha].value).split(":")
                elem.send_keys(horas1[0])
                elem.send_keys(horas1[1])
                elem = Select(website.find_element_by_id("analysisDropList:1"))
                elem.select_by_value('230')
                elem = Select(website.find_element_by_id("manufacturerDropList:1"))
                elem.select_by_value('67')
                elem = Select(website.find_element_by_id("nameTestDropList:1"))
                elem.select_by_value('89')
                elem = Select(website.find_element_by_id("techniqueDropList:1-analysisDropList:1"))
                elem.select_by_value('633')
                elem = Select(website.find_element_by_id("resultDropList:1"))
                result = ws['C' + linha].value
                if result == "P" or result == "p":
                    elem.select_by_value('70')
                elif result == "N" or result == "n":
                    elem.select_by_value('71')
                elem = website.find_element_by_id("obs_resultado_form:1")
                elem.send_keys(ws['E'+linha].value)
                elem = website.find_element_by_id("data_validacao_form:1")
                elem.send_keys(today[0])
                elem.send_keys(today[1])
                elem.send_keys(today[2])
                horas2 = horas1
                horas2[1] = str(int(horas1[1]) + 15)
                if int(horas2[1]) >= 60:
                    horas2[0] = str(int(horas2[0]) + 1)
                    horas2[1] = str(int(horas2[1]) - 60)
                elem.send_keys(horas2[0])
                elem.send_keys(horas2[1])
                elem = Select(website.find_element_by_id("agentDropList:1"))
                elem.select_by_value('259')
                time.sleep(10000000)
                elem = website.find_element_by_id("nextbtn")
                elem.click()
                WebDriverWait(website, 10).until(ec.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/"
                                                                                       "div/div[2]/button[2]"))).click()
                alert = website.switch_to.alert
                alert.accept()
                Label(root, text="Foram registados %d testes." % contador, bg='grey').place(x=0, y=44)
                ws['G' + linha].value = "Sim"
            else:
                ws['G' + linha].value = "Não"
        except:
            mensage(linha)
            slide()
            ws['G' + linha].value = "Não"
            website.get('https://sinave.min-saude.pt/SINAVE.MIN-SAUDE/Index')
        if str(ws['C' + str(int(linha)+1)].value) not in "-NPnp":
            break
    total = int(linha)
    today = date.today()
    today = today.strftime("%d/%m/%Y")
    today = today.replace('/', '-')
    website.get("https://sinave.min-saude.pt/SINAVE.MIN-SAUDE/SearchServlet")
    elem = Select(website.find_element_by_id('id_dis'))
    elem.select_by_value('62')
    elem = website.find_element_by_id('datini')
    elem.send_keys(str(today))
    elem = website.find_element_by_id('datfim')
    elem.send_keys(str(today))
    elem = website.find_element_by_id('btPesq')
    elem.click()
    elem = Select(website.find_element_by_name("example2_length"))
    elem.select_by_value('100')
    count = 1
    for i in range(total):
        linha = str(total - i)
        if str(ws['G'+linha].value) == "Sim":
            elem = website.find_element_by_xpath('//*[@id="example2"]/tbody/tr[%d]/td[2]' % count)
            ws['D' + linha].value = str(elem.text)
            count += 1
    Label(root, text="Foram registados %d testes normais com sucesso" % contador, fg='#22A044',
          font=('Nunito', 12, "bold")).place(x=300, y=410, anchor="center")
    wb.save('assets/Normais.xlsx')
    linha = 0
    wb.close()
    website.close()


def voith():
    numero = 10000
    contador = 0
    global linha
    wb = load_workbook('assets/Voith.xlsx')
    ws = wb.active
    today = date.today()
    today = today.strftime("%d/%m/%Y")
    today = today.split('/')
    website = webdriver.Chrome("assets/chromedriver.exe")
    website.get('https://sinave.min-saude.pt/SINAVE.MIN-SAUDE/login.html')
    elem = website.find_element_by_name("username")
    elem.send_keys('******')
    elem = website.find_element_by_name("password")
    elem.send_keys('*******')
    elem = website.find_element_by_name("button")
    elem.click()
    elem = website.find_element_by_name("chooselab")
    elem.click()
    for i in range(numero):
        linha = str(i + 2)
        try:
            if ws['C' + linha].value != "-":
                contador += 1
                notificacao = website.find_element_by_xpath('//*[@id="formTable"]/div/div[1]/div/div[1]/h3')
                notificacao = str(int(notificacao.text)+1)
                website.get('https://sinave.min-saude.pt/SINAVE.MIN-SAUDE/DiseaseNotificationPresc')
                elem = website.find_element_by_name("num_exame_form")
                elem.send_keys(notificacao)
                elem = website.find_element_by_name("presctyperadio")
                elem.click()
                elem = website.find_element_by_name("cod_local_colheita_form")
                elem.send_keys("*******")
                elem = website.find_element_by_name("codpostal_local_colheita_form")
                elem.send_keys("*******")
                elem = website.find_element_by_id("nextbtn")
                elem.click()
                elem = website.find_element_by_id("patient")
                utente = ws['B' + linha].value
                elem.send_keys(utente)
                elem = website.find_element_by_xpath('//*[@id="patient-box"]/div[1]/section[2]/label/button')
                elem.click()
                nome = website.find_element_by_name('nome_utente_form')
                nome.get_attribute("value")
                morada = website.find_element_by_name('morada_utente_form')
                morada.get_attribute("value")
                nascimento = website.find_element_by_name('utente_data_nasc_form')
                nascimento.get_attribute("value")
                elem = website.find_element_by_id("nextbtn")
                elem.click()
                elem = Select(website.find_element_by_id('diseaseDropList:1'))
                elem.select_by_value('62')
                elem = Select(website.find_element_by_id('productDropList:1'))
                elem.select_by_value('137')
                elem = website.find_element_by_id('data_colheita_form:1')
                elem.send_keys(today[0])
                elem.send_keys(today[1])
                elem.send_keys(today[2])
                horas1 = ["07", "30"]
                elem.send_keys(horas1[0])
                elem.send_keys(horas1[1])
                elem = Select(website.find_element_by_id("analysisDropList:1"))
                elem.select_by_value('230')
                elem = Select(website.find_element_by_id("manufacturerDropList:1"))
                elem.select_by_value('67')
                elem = Select(website.find_element_by_id("nameTestDropList:1"))
                elem.select_by_value('89')
                elem = Select(website.find_element_by_id("techniqueDropList:1-analysisDropList:1"))
                elem.select_by_value('633')
                elem = Select(website.find_element_by_id("resultDropList:1"))
                result = ws['C' + linha].value
                if result == "P" or result == "p":
                    elem.select_by_value('70')
                elif result == "N" or result == "n":
                    elem.select_by_value('71')
                elem = website.find_element_by_id("obs_resultado_form:1")
                elem.send_keys("R10")
                elem = website.find_element_by_id("data_validacao_form:1")
                elem.send_keys(today[0])
                elem.send_keys(today[1])
                elem.send_keys(today[2])
                horas2 = horas1
                horas2[1] = str(int(horas1[1]) + 15)
                if int(horas2[1]) >= 60:
                    horas2[0] = str(int(horas2[0]) + 1)
                    horas2[1] = str(int(horas2[1]) - 60)
                elem.send_keys(horas2[0])
                elem.send_keys(horas2[1])
                elem = Select(website.find_element_by_id("agentDropList:1"))
                elem.select_by_value('259')
                time.sleep(10000000)
                elem = website.find_element_by_id("nextbtn")
                elem.click()
                WebDriverWait(website, 10).until(ec.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/div"
                                                                                       "/div[2]/button[2]"))).click()
                alert = website.switch_to.alert
                alert.accept()
                Label(root, text="Foram registados %d testes." % contador, fg='purple', font=('Nunito', 12, "bold")).place(x=300, y=410, anchor="center")
        except:
            mensage(linha)
            slide()
            website.get('https://sinave.min-saude.pt/SINAVE.MIN-SAUDE/Index')
        if str(ws['C' + str(int(linha)+1)].value) not in "-NPnp":
            break
    Label(root, text="Foram registados %d testes da Voith com sucesso" % contador, fg='#22A044', font=('Nunito', 12, "bold")).place(x=300, y=410, anchor="center")
    wb.save('assets/Voith.xlsx')
    linha = 0
    wb.close()
    website.close()


linha = 0
msg = ""
root = Tk()
var = IntVar()
root.iconbitmap("assets/covid.ico")
root.title(string='Resgisto de Testes')
root.geometry('600x450')
root.resizable(False, False)
root.configure(bg='#EFEFEF')
img = (Image.open("assets/covid.png"))
new_image = ImageTk.PhotoImage(img)
Label(root, image=new_image).place(x=300, y=200, anchor="center")
Label(root, text=" Registo de Testes COVID-19 ", font=('Nunito', 22, 'bold'), fg="white", borderwidth=4, relief="solid",
      bg="#284B63").place(x=300, y=50, anchor="center")
buton = Button(root, text="   Testes VOITH   ", borderwidth=4, relief="solid", fg="white", font=('Nunito', 18, 'bold'),
               bg="#3C6E71", height=0, command=lambda: threading.Thread(target=voith).start())
buton.place(x=190, y=364, anchor="center")
buton = Button(root, text="  Testes Normais  ", font=('Nunito', 18, "bold"), borderwidth=4, fg="white", relief="solid",
               bg="#3C6E71", height=0, command=lambda: threading.Thread(target=normal).start())
buton.place(x=410, y=364, anchor="center")
root.mainloop()
